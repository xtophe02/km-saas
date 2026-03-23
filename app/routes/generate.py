"""Generation routes with async job pattern."""
import json
import time
import uuid
import threading
import logging
import calendar
import zipfile
from datetime import date, datetime
from io import BytesIO
from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, Response, RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from app.database import get_db, SessionLocal
from app.auth import get_current_user
from app.models import User, GenerationLog, CreditTransaction
from app.config import get_settings
from app.utils.generator import TripGenerator
from app.utils.report import KMReportGenerator

logger = logging.getLogger(__name__)

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
settings = get_settings()

# --- In-memory job store ---
_jobs = {}
_jobs_lock = threading.Lock()
_JOB_EXPIRY_SECONDS = 600  # 10 minutes


def _cleanup_expired_jobs():
    """Remove jobs older than expiry time."""
    now = time.time()
    with _jobs_lock:
        expired = [jid for jid, job in _jobs.items() if now - job["created"] > _JOB_EXPIRY_SECONDS]
        for jid in expired:
            del _jobs[jid]


def _get_active_job(user_id: int):
    """Find an active (processing) job for a user."""
    with _jobs_lock:
        for jid, job in _jobs.items():
            if job["user_id"] == user_id and job["status"] == "processing":
                return jid
    return None


def _run_generation(job_id: str, user_id: int, form: dict, km_limit: float = None):
    """Background thread: run TripGenerator and store result in job store."""
    try:
        with _jobs_lock:
            _jobs[job_id]["progress"] = "Building distance cache..."

        db = SessionLocal()
        try:
            generator = TripGenerator(
                db=db,
                office_address=form["departure_address"],
                target_kilometers=form["target_km"],
                table_name=form["code_sites_table"],
                num_days=form["num_days"],
            )

            with _jobs_lock:
                _jobs[job_id]["progress"] = "Generating trips..."

            trips = generator.generate_trips(
                month=form["month"], year=form["year"], km_limit=km_limit,
            )

            db.add(GenerationLog(
                user_id=user_id,
                consultant_name=form["consultant_name"],
                month=form["month"],
                year=form["year"],
                target_km=form["target_km"],
                num_days=form["num_days"],
                action="preview",
            ))
            db.commit()

            total_km = sum(t["total_distance"] for t in trips)
            trips_json = json.dumps(trips)

            with _jobs_lock:
                _jobs[job_id]["status"] = "done"
                _jobs[job_id]["progress"] = "Done"
                _jobs[job_id]["result"] = {
                    "trips": trips,
                    "total_km": total_km,
                    "trips_json": trips_json,
                }
        finally:
            db.close()

    except Exception as e:
        logger.error(f"Job {job_id} failed: {e}")
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["progress"] = str(e)


def _run_batch_generation(job_id: str, user_id: int, entries: list,
                          departure_address: str, code_sites_table: str,
                          month: int, year: int):
    """Background thread: generate PDFs for all batch entries, produce ZIP."""
    db = SessionLocal()
    try:
        total = len(entries)
        pdf_results = []
        report_gen = KMReportGenerator()
        report_date = date(year, month, 1)

        # Shared distance calculator: first TripGenerator builds the cache,
        # subsequent ones reuse it since distances are in the DB
        for idx, entry in enumerate(entries):
            with _jobs_lock:
                _jobs[job_id]["progress"] = f"Processing {idx + 1}/{total}: {entry['worker']}..."
                _jobs[job_id]["batch_done"] = idx

            target_km = entry["mileage_amount"]
            # Auto-calculate num_days: min(round(target_km / 200), workdays_in_month)
            workdays_in_month = _count_workdays(month, year)
            num_days = min(max(1, round(target_km / 200)), workdays_in_month)

            try:
                generator = TripGenerator(
                    db=db,
                    office_address=departure_address,
                    target_kilometers=target_km,
                    table_name=code_sites_table,
                    num_days=num_days,
                )
                trips = generator.generate_trips(month=month, year=year)

                if not trips:
                    logger.warning(f"Batch entry {idx}: no trips generated for {entry['worker']}")
                    pdf_results.append({
                        "worker": entry["worker"],
                        "status": "error",
                        "error": "No trips could be generated",
                    })
                    continue

                pdf_bytes = report_gen.generate_report(entry["worker"], report_date, trips)
                total_km = sum(t["total_distance"] for t in trips)
                filename = f"{entry['worker'].replace(' ', '_')}_{total_km:.1f}_KM_{month:02d}{year}.pdf"

                pdf_results.append({
                    "worker": entry["worker"],
                    "status": "ok",
                    "pdf_bytes": pdf_bytes,
                    "filename": filename,
                    "total_km": total_km,
                    "num_trips": len(trips),
                })

                # Log generation
                db.add(GenerationLog(
                    user_id=user_id,
                    consultant_name=entry["worker"],
                    month=month,
                    year=year,
                    target_km=target_km,
                    num_days=num_days,
                    action="batch_pdf",
                ))
                db.commit()

            except Exception as e:
                logger.error(f"Batch entry {idx} ({entry['worker']}) failed: {e}")
                pdf_results.append({
                    "worker": entry["worker"],
                    "status": "error",
                    "error": str(e),
                })

        # Build ZIP
        with _jobs_lock:
            _jobs[job_id]["progress"] = "Building ZIP archive..."

        zip_buffer = BytesIO()
        success_count = 0
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for result in pdf_results:
                if result["status"] == "ok":
                    zf.writestr(result["filename"], result["pdf_bytes"])
                    success_count += 1

        zip_bytes = zip_buffer.getvalue()
        zip_buffer.close()

        with _jobs_lock:
            _jobs[job_id]["status"] = "done"
            _jobs[job_id]["progress"] = "Done"
            _jobs[job_id]["batch_done"] = total
            _jobs[job_id]["result"] = {
                "zip_bytes": zip_bytes,
                "total_entries": total,
                "success_count": success_count,
                "error_count": total - success_count,
                "results": [
                    {k: v for k, v in r.items() if k != "pdf_bytes"}
                    for r in pdf_results
                ],
            }

    except Exception as e:
        logger.error(f"Batch job {job_id} failed: {e}")
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["progress"] = str(e)
    finally:
        db.close()


def _count_workdays(month: int, year: int) -> int:
    """Count weekday days in a month (simplified, excludes holidays)."""
    _, last_day = calendar.monthrange(year, month)
    count = 0
    for day in range(1, last_day + 1):
        d = date(year, month, day)
        if d.weekday() < 5:
            count += 1
    return count


# ============================================================
# Single generation endpoints (existing)
# ============================================================

@router.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    # Check for active job
    active_job_id = _get_active_job(user.id)

    logs = (
        db.query(GenerationLog)
        .filter(GenerationLog.user_id == user.id)
        .order_by(GenerationLog.created_at.desc())
        .limit(20)
        .all()
    )
    return templates.TemplateResponse("dashboard.html", {
        "request": request, "user": user, "logs": logs,
        "active_job_id": active_job_id,
    })


@router.get("/generate", response_class=HTMLResponse)
async def generate_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    # Check for active job — redirect back to processing page
    active_job_id = _get_active_job(user.id)
    if active_job_id:
        return templates.TemplateResponse("processing.html", {
            "request": request, "user": user, "job_id": active_job_id,
        })

    now = date.today()
    return templates.TemplateResponse("generate.html", {
        "request": request, "user": user,
        "current_month": now.month, "current_year": now.year,
        "max_km_per_day": settings.max_km_per_day,
        "preview": None, "form": {},
    })


@router.post("/generate/preview", response_class=HTMLResponse)
async def generate_preview(
    request: Request,
    consultant_name: str = Form(...),
    month: int = Form(...),
    year: int = Form(...),
    target_km: float = Form(...),
    num_days: int = Form(...),
    departure_address: str = Form(...),
    db: Session = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    # If there's already an active job, redirect to it
    active_job_id = _get_active_job(user.id)
    if active_job_id:
        return templates.TemplateResponse("processing.html", {
            "request": request, "user": user, "job_id": active_job_id,
        })

    form = {
        "consultant_name": consultant_name, "month": month, "year": year,
        "target_km": target_km, "num_days": num_days,
        "departure_address": departure_address,
    }
    errors = _validate_form(form)

    if errors:
        return templates.TemplateResponse("generate.html", {
            "request": request, "user": user,
            "current_month": month, "current_year": year,
            "max_km_per_day": settings.max_km_per_day,
            "preview": None, "form": form, "errors": errors,
        })

    km_limit = settings.free_tier_km_limit if user.credits <= 0 else None
    is_free_tier = user.credits <= 0

    _cleanup_expired_jobs()

    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "processing",
            "progress": "Starting...",
            "created": time.time(),
            "result": None,
            "form": form,
            "user_id": user.id,
            "is_free_tier": is_free_tier,
        }

    form["code_sites_table"] = user.code_sites_table

    thread = threading.Thread(
        target=_run_generation,
        args=(job_id, user.id, form, km_limit),
        daemon=True,
    )
    thread.start()

    return templates.TemplateResponse("processing.html", {
        "request": request, "user": user, "job_id": job_id,
    })


@router.get("/generate/single/status/{job_id}")
async def generate_status(job_id: str, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"status": "error", "progress": "Not authenticated"}, status_code=401)

    with _jobs_lock:
        job = _jobs.get(job_id)

    if not job:
        return JSONResponse({"status": "error", "progress": "Job not found or expired"})

    if job["user_id"] != user.id:
        return JSONResponse({"status": "error", "progress": "Unauthorized"}, status_code=403)

    resp = {
        "status": job["status"],
        "progress": job["progress"],
    }
    # Include batch progress info if available
    if "batch_done" in job:
        resp["batch_done"] = job["batch_done"]
    if "batch_total" in job:
        resp["batch_total"] = job["batch_total"]

    return JSONResponse(resp)


@router.get("/generate/single/result/{job_id}", response_class=HTMLResponse)
async def generate_result(job_id: str, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    with _jobs_lock:
        job = _jobs.get(job_id)

    if not job or job["user_id"] != user.id:
        return RedirectResponse("/generate", status_code=303)

    if job["status"] == "error":
        form = job["form"]
        return templates.TemplateResponse("generate.html", {
            "request": request, "user": user,
            "current_month": form.get("month", date.today().month),
            "current_year": form.get("year", date.today().year),
            "max_km_per_day": settings.max_km_per_day,
            "preview": None, "form": form,
            "errors": [job["progress"]],
        })

    if job["status"] != "done":
        return templates.TemplateResponse("processing.html", {
            "request": request, "user": user, "job_id": job_id,
        })

    form = job["form"]
    result = job["result"]

    return templates.TemplateResponse("generate.html", {
        "request": request, "user": user,
        "current_month": form["month"], "current_year": form["year"],
        "max_km_per_day": settings.max_km_per_day,
        "preview": {
            "trips": result["trips"],
            "total_km": result["total_km"],
            "trips_json": result["trips_json"],
        },
        "form": form, "is_free_tier": job.get("is_free_tier", False),
    })


@router.post("/generate/pdf")
async def generate_pdf(
    request: Request,
    consultant_name: str = Form(...),
    month: int = Form(...),
    year: int = Form(...),
    target_km: float = Form(...),
    num_days: int = Form(...),
    departure_address: str = Form(...),
    trips_json: str = Form(""),
    db: Session = Depends(get_db),
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    is_free_tier = user.credits <= 0

    try:
        # Use trips from the preview (same data the user saw)
        if trips_json:
            trips = json.loads(trips_json)
        else:
            # Fallback: regenerate (shouldn't normally happen)
            km_limit = settings.free_tier_km_limit if is_free_tier else None
            generator = TripGenerator(
                db=db, office_address=departure_address,
                target_kilometers=target_km, table_name=user.code_sites_table,
                num_days=num_days,
            )
            trips = generator.generate_trips(month=month, year=year, km_limit=km_limit)

        if not trips:
            return RedirectResponse("/generate", status_code=303)

        report = KMReportGenerator()
        report_date = date(year, month, 1)
        pdf_bytes = report.generate_report(consultant_name, report_date, trips)

        # Only deduct credit if not free tier
        if not is_free_tier:
            user.credits -= 1
            db.add(CreditTransaction(
                user_id=user.id, amount=-1, description="PDF generation",
            ))
        db.add(GenerationLog(
            user_id=user.id, consultant_name=consultant_name,
            month=month, year=year, target_km=target_km,
            num_days=num_days, action="pdf" if not is_free_tier else "free_pdf",
        ))
        db.commit()

        total_distance = sum(t["total_distance"] for t in trips)
        filename = f"{consultant_name.replace(' ', '_')}_{total_distance:.1f}_KM_{month:02d}{year}.pdf"

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as e:
        logger.error(f"PDF generation error: {e}")
        return RedirectResponse("/generate", status_code=303)


# ============================================================
# Batch generation endpoints
# ============================================================

@router.get("/generate/batch", response_class=HTMLResponse)
async def batch_page(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not user.batch_enabled:
        return RedirectResponse("/generate", status_code=303)

    # Check for active batch job
    active_job_id = _get_active_job(user.id)
    if active_job_id:
        with _jobs_lock:
            job = _jobs.get(active_job_id, {})
        if job.get("job_type") == "batch":
            return templates.TemplateResponse("batch_processing.html", {
                "request": request, "user": user, "job_id": active_job_id,
                "batch_total": job.get("batch_total", 0),
            })

    now = date.today()
    return templates.TemplateResponse("batch.html", {
        "request": request, "user": user,
        "current_month": now.month, "current_year": now.year,
    })


@router.get("/generate/batch/template")
async def batch_template(request: Request, db: Session = Depends(get_db)):
    """Download a template xlsx with the correct columns."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Batch KM"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4E3BC2", end_color="4E3BC2", fill_type="solid")

    headers = ["Worker", "Effective date", "Mileage Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example rows
    ws.cell(row=2, column=1, value="John Doe")
    ws.cell(row=2, column=2, value="03/2026")
    ws.cell(row=2, column=3, value=1500)

    ws.cell(row=3, column=1, value="Jane Smith")
    ws.cell(row=3, column=2, value="03/2026")
    ws.cell(row=3, column=3, value=2200)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="batch_km_template.xlsx"'},
    )


@router.post("/generate/batch/upload", response_class=HTMLResponse)
async def batch_upload(
    request: Request,
    file: UploadFile = File(...),
    departure_address: str = Form(""),
    db: Session = Depends(get_db),
):
    """Parse uploaded xlsx and show preview."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not user.batch_enabled:
        return RedirectResponse("/generate", status_code=303)

    departure_address = departure_address.strip() or user.default_address or ""

    from openpyxl import load_workbook

    errors = []
    entries = []

    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        # Check headers
        headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
        required = {"worker", "effective date", "mileage amount"}

        col_map = {}
        for idx, h in enumerate(headers):
            if "worker" in h:
                col_map["worker"] = idx
            elif "date" in h or "effective" in h:
                col_map["date"] = idx
            elif "mileage" in h or "amount" in h or "km" in h:
                col_map["mileage"] = idx

        if len(col_map) < 3:
            missing = []
            if "worker" not in col_map:
                missing.append("Worker")
            if "date" not in col_map:
                missing.append("Effective date")
            if "mileage" not in col_map:
                missing.append("Mileage Amount")
            errors.append(f"Missing columns: {', '.join(missing)}. Found headers: {headers}")

            now = date.today()
            return templates.TemplateResponse("batch.html", {
                "request": request, "user": user, "errors": errors,
                "current_month": now.month, "current_year": now.year,
            })

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_values = [cell.value for cell in row]

            worker = row_values[col_map["worker"]]
            date_val = row_values[col_map["date"]]
            mileage_val = row_values[col_map["mileage"]]

            # Skip empty rows
            if not worker and not date_val and not mileage_val:
                continue

            row_errors = []

            # Validate worker
            if not worker or not str(worker).strip():
                row_errors.append(f"Row {row_idx}: Worker name is empty")
                worker = ""
            else:
                worker = str(worker).strip()

            # Parse date: accept "MM/YYYY", "YYYY-MM", or datetime
            month = None
            year = None
            if date_val:
                date_str = str(date_val).strip()
                try:
                    if "/" in date_str:
                        parts = date_str.split("/")
                        month = int(parts[0])
                        year = int(parts[1])
                    elif "-" in date_str:
                        parts = date_str.split("-")
                        if len(parts[0]) == 4:
                            year = int(parts[0])
                            month = int(parts[1])
                        else:
                            month = int(parts[0])
                            year = int(parts[1])
                    elif hasattr(date_val, "month"):
                        month = date_val.month
                        year = date_val.year
                    else:
                        row_errors.append(f"Row {row_idx}: Invalid date format '{date_str}' (use MM/YYYY)")
                except (ValueError, IndexError):
                    row_errors.append(f"Row {row_idx}: Invalid date format '{date_str}' (use MM/YYYY)")
            else:
                row_errors.append(f"Row {row_idx}: Effective date is empty")

            if month is not None and (month < 1 or month > 12):
                row_errors.append(f"Row {row_idx}: Invalid month {month}")
                month = None
            if year is not None and (year < 2020 or year > 2030):
                row_errors.append(f"Row {row_idx}: Invalid year {year}")
                year = None

            # Validate mileage
            km = 0
            if mileage_val is not None:
                try:
                    km = float(mileage_val)
                    if km <= 0:
                        row_errors.append(f"Row {row_idx}: Mileage must be > 0 (got {km})")
                except (ValueError, TypeError):
                    row_errors.append(f"Row {row_idx}: Invalid mileage value '{mileage_val}'")
            else:
                row_errors.append(f"Row {row_idx}: Mileage Amount is empty")

            errors.extend(row_errors)

            if not row_errors and month and year:
                entries.append({
                    "worker": worker,
                    "month": month,
                    "year": year,
                    "mileage_amount": km,
                    "date_str": f"{month:02d}/{year}",
                })

        wb.close()

    except Exception as e:
        errors.append(f"Failed to read file: {str(e)}")

    if not entries and not errors:
        errors.append("No data rows found in the file")

    # Store validated entries in job store for the run step
    _cleanup_expired_jobs()
    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "preview",
            "progress": "",
            "created": time.time(),
            "result": None,
            "user_id": user.id,
            "job_type": "batch_preview",
            "entries": entries,
            "departure_address": departure_address,
        }

    credits_needed = len(entries)
    has_enough = user.credits >= credits_needed

    return templates.TemplateResponse("batch_preview.html", {
        "request": request, "user": user,
        "entries": entries, "errors": errors,
        "credits_needed": credits_needed,
        "has_enough": has_enough,
        "job_id": job_id,
        "departure_address": departure_address,
    })


@router.post("/generate/batch/run")
async def batch_run(
    request: Request,
    db: Session = Depends(get_db),
):
    """Start batch generation from validated preview data."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)
    if not user.batch_enabled:
        return RedirectResponse("/generate", status_code=303)

    form = await request.form()
    preview_job_id = form.get("job_id", "")

    with _jobs_lock:
        preview_job = _jobs.get(preview_job_id)

    if not preview_job or preview_job["user_id"] != user.id or preview_job.get("job_type") != "batch_preview":
        return RedirectResponse("/generate/batch", status_code=303)

    entries = preview_job.get("entries", [])
    if not entries:
        return RedirectResponse("/generate/batch", status_code=303)

    credits_needed = len(entries)

    # Re-check credits (user might have spent them)
    db.refresh(user)
    if user.credits < credits_needed:
        return templates.TemplateResponse("batch_preview.html", {
            "request": request, "user": user,
            "entries": entries, "errors": [f"Not enough credits. Need {credits_needed}, have {user.credits}."],
            "credits_needed": credits_needed,
            "has_enough": False,
            "job_id": preview_job_id,
        })

    # Deduct all credits at once
    user.credits -= credits_needed
    db.add(CreditTransaction(
        user_id=user.id, amount=-credits_needed,
        description=f"Batch generation: {credits_needed} PDFs",
    ))
    db.commit()

    # Remove preview job, create processing job
    with _jobs_lock:
        if preview_job_id in _jobs:
            del _jobs[preview_job_id]

    job_id = str(uuid.uuid4())
    with _jobs_lock:
        _jobs[job_id] = {
            "status": "processing",
            "progress": "Starting batch generation...",
            "created": time.time(),
            "result": None,
            "user_id": user.id,
            "job_type": "batch",
            "batch_total": len(entries),
            "batch_done": 0,
        }

    # All entries share the same departure address and code_sites_table
    departure_address = preview_job.get("departure_address") or user.default_address
    code_sites_table = user.code_sites_table

    # Group entries by month/year for the background thread
    # (all entries could have different months, but the generator handles one at a time)
    # For simplicity, pass the first entry's month/year as default,
    # but the actual implementation uses each entry's own month/year
    thread = threading.Thread(
        target=_run_batch_generation_multi,
        args=(job_id, user.id, entries, departure_address, code_sites_table),
        daemon=True,
    )
    thread.start()

    return templates.TemplateResponse("batch_processing.html", {
        "request": request, "user": user, "job_id": job_id,
        "batch_total": len(entries),
    })


def _run_batch_generation_multi(job_id: str, user_id: int, entries: list,
                                departure_address: str, code_sites_table: str):
    """Background thread: generate PDFs for batch entries (each may have different month/year)."""
    db = SessionLocal()
    try:
        total = len(entries)
        pdf_results = []
        report_gen = KMReportGenerator()

        for idx, entry in enumerate(entries):
            with _jobs_lock:
                _jobs[job_id]["progress"] = f"Processing {idx + 1}/{total}: {entry['worker']}..."
                _jobs[job_id]["batch_done"] = idx

            month = entry["month"]
            year = entry["year"]
            target_km = entry["mileage_amount"]
            report_date = date(year, month, 1)

            # Auto-calculate num_days
            workdays_in_month = _count_workdays(month, year)
            num_days = min(max(1, round(target_km / 200)), workdays_in_month)

            try:
                generator = TripGenerator(
                    db=db,
                    office_address=departure_address,
                    target_kilometers=target_km,
                    table_name=code_sites_table,
                    num_days=num_days,
                )
                trips = generator.generate_trips(month=month, year=year)

                if not trips:
                    logger.warning(f"Batch entry {idx}: no trips for {entry['worker']}")
                    pdf_results.append({
                        "worker": entry["worker"],
                        "date_str": entry["date_str"],
                        "status": "error",
                        "error": "No trips could be generated",
                    })
                    continue

                pdf_bytes = report_gen.generate_report(entry["worker"], report_date, trips)
                total_km = sum(t["total_distance"] for t in trips)
                filename = f"{entry['worker'].replace(' ', '_')}_{total_km:.1f}_KM_{month:02d}{year}.pdf"

                pdf_results.append({
                    "worker": entry["worker"],
                    "date_str": entry["date_str"],
                    "status": "ok",
                    "pdf_bytes": pdf_bytes,
                    "filename": filename,
                    "total_km": total_km,
                    "num_trips": len(trips),
                })

                db.add(GenerationLog(
                    user_id=user_id,
                    consultant_name=entry["worker"],
                    month=month,
                    year=year,
                    target_km=target_km,
                    num_days=num_days,
                    action="batch_pdf",
                ))
                db.commit()

            except Exception as e:
                logger.error(f"Batch entry {idx} ({entry['worker']}) failed: {e}")
                pdf_results.append({
                    "worker": entry["worker"],
                    "date_str": entry["date_str"],
                    "status": "error",
                    "error": str(e),
                })

        # Build ZIP
        with _jobs_lock:
            _jobs[job_id]["progress"] = "Building ZIP archive..."

        zip_buffer = BytesIO()
        success_count = 0
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for result in pdf_results:
                if result["status"] == "ok":
                    zf.writestr(result["filename"], result["pdf_bytes"])
                    success_count += 1

        zip_bytes = zip_buffer.getvalue()
        zip_buffer.close()

        with _jobs_lock:
            _jobs[job_id]["status"] = "done"
            _jobs[job_id]["progress"] = "Done"
            _jobs[job_id]["batch_done"] = total
            _jobs[job_id]["result"] = {
                "zip_bytes": zip_bytes,
                "total_entries": total,
                "success_count": success_count,
                "error_count": total - success_count,
                "results": [
                    {k: v for k, v in r.items() if k != "pdf_bytes"}
                    for r in pdf_results
                ],
            }

    except Exception as e:
        logger.error(f"Batch job {job_id} failed: {e}")
        with _jobs_lock:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["progress"] = str(e)
    finally:
        db.close()


@router.get("/generate/batch/status/{job_id}")
async def batch_status(job_id: str, request: Request, db: Session = Depends(get_db)):
    """JSON status endpoint for batch jobs."""
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"status": "error", "progress": "Not authenticated"}, status_code=401)

    with _jobs_lock:
        job = _jobs.get(job_id)

    if not job:
        return JSONResponse({"status": "error", "progress": "Job not found or expired"})

    if job["user_id"] != user.id:
        return JSONResponse({"status": "error", "progress": "Unauthorized"}, status_code=403)

    return JSONResponse({
        "status": job["status"],
        "progress": job["progress"],
        "batch_done": job.get("batch_done", 0),
        "batch_total": job.get("batch_total", 0),
    })


@router.get("/generate/batch/result/{job_id}", response_class=HTMLResponse)
async def batch_result(job_id: str, request: Request, db: Session = Depends(get_db)):
    """Show batch results page."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    with _jobs_lock:
        job = _jobs.get(job_id)

    if not job or job["user_id"] != user.id:
        return RedirectResponse("/generate/batch", status_code=303)

    if job["status"] == "error":
        return templates.TemplateResponse("batch.html", {
            "request": request, "user": user,
            "errors": [job["progress"]],
            "current_month": date.today().month, "current_year": date.today().year,
        })

    if job["status"] != "done":
        return templates.TemplateResponse("batch_processing.html", {
            "request": request, "user": user, "job_id": job_id,
            "batch_total": job.get("batch_total", 0),
        })

    result = job["result"]
    return templates.TemplateResponse("batch_done.html", {
        "request": request, "user": user, "job_id": job_id,
        "total_entries": result["total_entries"],
        "success_count": result["success_count"],
        "error_count": result["error_count"],
        "results": result["results"],
    })


@router.get("/generate/batch/download/{job_id}")
async def batch_download(job_id: str, request: Request, db: Session = Depends(get_db)):
    """Download the ZIP of generated PDFs."""
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=303)

    with _jobs_lock:
        job = _jobs.get(job_id)

    if not job or job["user_id"] != user.id or job["status"] != "done":
        return RedirectResponse("/generate/batch", status_code=303)

    zip_bytes = job["result"]["zip_bytes"]
    return Response(
        content=zip_bytes,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="batch_km_reports.zip"'},
    )


def _validate_form(form: dict) -> list:
    errors = []
    settings_obj = get_settings()

    if not form.get("consultant_name", "").strip():
        errors.append("Consultant name is required")
    if not form.get("departure_address", "").strip():
        errors.append("Departure address is required")

    km = form.get("target_km", 0)
    days = form.get("num_days", 0)

    if km <= 0:
        errors.append("Total kilometers must be positive")
    if days <= 0 or days > 31:
        errors.append("Number of days must be between 1 and 31")
    if km > 0 and days > 0 and km > days * settings_obj.max_km_per_day:
        errors.append(f"Maximum {settings_obj.max_km_per_day:.0f} km/day ({days} days = {days * settings_obj.max_km_per_day:.0f} km max)")

    month = form.get("month", 0)
    year = form.get("year", 0)
    if month < 1 or month > 12:
        errors.append("Invalid month")
    if year < 2020 or year > 2030:
        errors.append("Year must be between 2020 and 2030")

    return errors
